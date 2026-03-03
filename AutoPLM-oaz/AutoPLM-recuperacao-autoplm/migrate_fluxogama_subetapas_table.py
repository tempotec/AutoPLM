"""
Seed e migration da tabela fluxogama_subetapas.
Lê subetapas_25022026_171109.xlsx (coluna WSID + nome) e popula o banco.

Executar: python migrate_fluxogama_subetapas_table.py
"""
import os
import sys

# Load env
env_path = os.path.join(os.path.dirname(__file__), '.env.local')
if os.path.exists(env_path):
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if '=' in line and not line.startswith('#'):
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip())

DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    print("❌ DATABASE_URL não configurada.")
    sys.exit(1)


def normalize_wsid(x: str) -> str:
    """Normaliza WSID: string, strip, "14.0" → "14"."""
    s = str(x or '').strip()
    if not s or s == 'None':
        return ''
    try:
        if '.' in s:
            s = str(int(float(s)))
    except (ValueError, TypeError):
        pass
    return s


# Known subetapas — base inicial (pode ser sobrescrita pelo xlsx)
KNOWN_SUBETAPAS = [
    ('14', 'Ficha do desenvolvimento'),
    ('54', 'Entrada do mix Souq'),
    ('55', 'Entrada do mix TSM'),
    ('70', 'Tabela de Medidas'),
    ('11', 'Fluxogama'),
    ('51', 'Bases de Etiquetas e Tag'),
    ('52', 'Kit de Aviamentos'),
    ('15', 'Cotação/Escolha com Fornecedor'),
    ('16', 'Montagem da Coleção'),
    ('17', 'Prova de Roupa/Lacre'),
    ('18', 'Envio comentários p/ Fornecedor'),
    ('19', 'Aprovação do Produto Lacre/Ppsample'),
    ('69', 'Desenhos à fazer'),
    ('57', 'Treinamento'),
    ('1',  'Cadastros auxiliares'),
    ('2',  'Cadastros auxiliares (detalhe)'),
    ('3',  'Estilo importado'),
]


def load_from_xlsx():
    """Tenta carregar subetapas do xlsx exportado do Fluxogama."""
    xlsx_path = os.path.join(os.path.dirname(__file__),
                             r'C:\Users\USER\Downloads\oaz\subetapas_25022026_171109 (1).xlsx')
    if not os.path.exists(xlsx_path):
        # Try relative
        xlsx_path = r'C:\Users\USER\Downloads\oaz\subetapas_25022026_171109 (1).xlsx'
    if not os.path.exists(xlsx_path):
        print("ℹ️  xlsx não encontrado, usando lista padrão.")
        return None

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip().lower() for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]

        # Find WSID and name columns
        wsid_col = next((i for i, h in enumerate(headers) if 'wsid' in h), None)
        name_col = next((i for i, h in enumerate(headers) if any(k in h for k in ['estrutura', 'nome', 'name', 'processo'])), None)

        if wsid_col is None:
            print(f"⚠️  Coluna WSID não encontrada no xlsx. Headers: {headers}")
            return None

        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[wsid_col]:
                continue
            wsid = normalize_wsid(row[wsid_col])
            nome = str(row[name_col] if name_col is not None else '').strip() or f'Subetapa {wsid}'
            if wsid:
                results.append((wsid, nome))

        print(f"✅ xlsx carregado: {len(results)} subetapas encontradas.")
        return results
    except Exception as e:
        print(f"⚠️  Erro ao ler xlsx: {e}. Usando lista padrão.")
        return None


def migrate():
    import psycopg2
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    # Create table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS fluxogama_subetapas (
            id SERIAL PRIMARY KEY,
            wsid VARCHAR(20) UNIQUE NOT NULL,
            nome VARCHAR(100) NOT NULL,
            ativo BOOLEAN NOT NULL DEFAULT TRUE,
            colecao_wsid VARCHAR(20)
        )
    """)
    conn.commit()
    print("✅ Tabela 'fluxogama_subetapas' criada/verificada.")

    # Load data
    subetapas = load_from_xlsx() or KNOWN_SUBETAPAS

    inserted = 0
    skipped = 0
    for wsid, nome in subetapas:
        wsid = normalize_wsid(wsid)
        if not wsid:
            continue
        try:
            cur.execute("""
                INSERT INTO fluxogama_subetapas (wsid, nome, ativo, colecao_wsid)
                VALUES (%s, %s, TRUE, NULL)
                ON CONFLICT (wsid) DO UPDATE SET nome = EXCLUDED.nome
            """, (wsid, nome))
            conn.commit()
            inserted += 1
        except Exception as e:
            conn.rollback()
            print(f"  ⚠️  wsid={wsid}: {e}")
            skipped += 1

    print(f"✅ Seed: {inserted} inseridas/atualizadas, {skipped} ignoradas.")
    conn.close()


if __name__ == '__main__':
    migrate()
